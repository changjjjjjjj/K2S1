import cv2 
import numpy as np
from matplotlib import pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color


class test:
    def __init__(self, path_origin, path_signed, path_guide):
        
        img_origin = cv2.imread(path_origin,cv2.IMREAD_COLOR)
        img_signed = cv2.imread(path_signed,cv2.IMREAD_COLOR)
        img_guide = cv2.imread(path_guide,cv2.IMREAD_COLOR)
        
        gray_origin = cv2.cvtColor(img_origin, cv2.COLOR_BGR2GRAY)
        gray_signed = cv2.cvtColor(img_signed, cv2.COLOR_BGR2GRAY)
        gray_guide = cv2.cvtColor(img_guide, cv2.COLOR_BGR2GRAY)
        
        ret, self.binary_origin = cv2.threshold(gray_origin, 127, 255, cv2.THRESH_BINARY)
        ret, self.binary_signed = cv2.threshold(gray_signed, 127, 255, cv2.THRESH_BINARY)
        ret, self.binary_guide = cv2.threshold(gray_guide, 127, 255, cv2.THRESH_BINARY)

        # init data bundle
        self.bundle_count = 0
        self.bundle_data = np.array([[0,0,0]])

    def cut(self):
        contours, hierachy = cv2.findContours(self.binary_guide, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
        
        self.bundle_count = len(contours)
        
        for i in range(self.bundle_count):
            x, y, w, h = cv2.boundingRect(contours[i])
            cut_origin = self.binary_origin[y:y+h, x:x+w]
            cut_signed = self.binary_signed[y:y+h, x:x+w]
            
            check = self.check(cut_origin, cut_signed)
            cut_data = np.array([[y+(h/2), x+(w/2), check]])
            # cut_data format : [y, x, check], check: 0(false), 1(true)

            self.bundle_data = np.append(self.bundle_data, cut_data, axis=0)

        self.bundle_data = self.bundle_data[1:]
        
        # img_guide 에서 contour 추출
        # contour에서 boundingRect 를 이용해 각 영역을 잘라내
        # bundle_origin, bundle_signed, bundle_position 값을 저장
        # guide [white: target, black: background]

    def check(self, origin, signed):
        value_origin = sum(sum(origin))
        value_signed = sum(sum(signed))
        ratio = value_signed / value_origin
        
        if ratio <= 1.003 : #1.01
            return 0
        else :
            print("Origin: %d, Signed: %d, Ratio: %.3f" %(value_origin, value_signed, ratio))
            cv2.imshow('origin', origin)
            cv2.imshow('signed', signed)
            cv2.waitKey(0)
            cv2.destroyWindow('origin')
            cv2.destroyWindow('signed')
            return 1

        # 알고리즘 1 : 영역의 오염를 측정하여 높으면 체크한 영역으로 감지 (그 한계를 잘 설정하는것이 관건)
        # 알고리즘 2 : 영역의 

    def save(self):
        print ("save")
        for i in range(self.bundle_count):
            data = self.bundle_data[i]
            print("Check: %1d, Position: [%.1f, %.1f]" %(data[2], data[0], data[1]))
    
    def write_to_excel(self):
        wb = Workbook()
        sheet1 = wb.active
        file_name = 'sample.xlsx'
        sheet1.title = '배뇨일지'

        sheet1.cell(row = 1, column = 1).value = '번호||정도'
        sheet1.cell(row = 1, column = 1).fill = PatternFill("solid", fgColor = "C0C0C0")
        sheet1.cell(row = 1, column = 7).value = '표시번호'
        sheet1.cell(row = 1, column = 7).fill = PatternFill("solid", fgColor = "C0C0C0")

        check_count = 0

        for i in range(1,6):
            sheet1.cell(row = 1, column = i+1).value = i-1
            sheet1.cell(row = 1, column = i+1).fill = PatternFill("solid", fgColor = "C0C0C0")
        
            
        for row_index in range(0, self.bundle_count):
    
            if row_index%5 == 0:
                sheet1.cell(row = row_index//5 + 2, column = 1).value = row_index//5 + 1
                sheet1.cell(row = row_index//5 + 2, column = 1).fill = PatternFill("solid", fgColor="C0C0C0")
                check_count = 0

            if (self.bundle_data[-(row_index+1)][2]) == 1:
                sheet1.cell(row = (row_index//5 + 2), column = (row_index%5) + 2).value = self.bundle_data[-(row_index+1)][2]
                sheet1.cell(row = (row_index//5 + 2), column = (row_index%5) + 2).fill = PatternFill("solid", fgColor = "FFFF00")
                sheet1.cell(row = (row_index//5 + 2), column = 7).value = row_index%5
                check_count += 1

                if check_count >=2:
                    sheet1.cell(row = (row_index//5 + 2), column = 7).fill = PatternFill("solid", fgColor = "FF0000")
                else :
                    sheet1.cell(row = row_index//5 + 2, column = 7).fill = PatternFill("solid", fgColor="00FFFF")
                
            else :
                sheet1.cell(row = (row_index//5 + 2), column = (row_index%5) + 2).value = self.bundle_data[-(row_index+1)][2]

        wb.save(filename = file_name)


if __name__ == '__main__':

    filepath1 = r"o1.jpg"
    filepath2 = r"i5.jpg"
    filepath3 = r"g1.jpg"

    test = test(filepath1, filepath2, filepath3)
    test.cut()
    test.save()
    test.write_to_excel()
